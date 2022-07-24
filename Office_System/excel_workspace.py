from openpyxl import load_workbook, drawing
import os
from PIL import Image
import global_var as gv

error_write_price_step = []


def create_path(path):
    if not os.path.isdir(path):
        os.mkdir(path)


def edit_price_file(sizes, code):
    err = "價格表出錯或是沒有關"
    try:
        base_path = "../source/"
        wb = load_workbook(f"{base_path}/source/price.xlsx")
        sheet = wb["price"]

        for i in range(len(sizes)):
            sheet.cell(column=2 + i, row=1).value = sizes[i]

        err = "產品編號出錯"
        for i in range(len(code[0])):
            sheet.cell(column=1, row=2 + i).value = code[0][i]

        wb.save(f"{base_path}價格表.xlsx")
        err = ""
        return err
    except:
        return err


# ------------------------------------------------------------------------------
def input_bar(sizes):
    base_path = "../source/"
    wb = load_workbook(f"{base_path}source/basic_order_{len(sizes)}.xlsx")
    sheet = wb["order"]
    bar_list = ["Price", "Lots", "QTY", "Subtotal"]
    for i in range(len(sizes)):
        for j in range(len(bar_list)):
            sheet.cell(column=3 + i * 4 + j, row=10).value = f"({sizes[i]})\n{bar_list[j]}"

    wb.save(f"{base_path}source/temp.xlsx")


# -----------------------------------------------------------------------------------
def input_code(code):
    base_path = "../source/"
    wb = load_workbook(f"{base_path}source/temp.xlsx")
    sheet = wb["order"]
    for i in range(len(code)):
        sheet.cell(column=2, row=11 + i).value = code[i]
    wb.save(f"{base_path}source/temp.xlsx")


def hide_col(n):
    base_path = "../source/"
    wb = load_workbook(f"{base_path}source/temp.xlsx")
    sheet = wb["order"]
    """
    for i in range(n):
        sheet.insert_rows(12)
    """
    # sheet.cell(row=15, column=4).border = copy(sheet.cell(row=11, column=4).border)
    sheet.row_dimensions.group(n + 11 + 1, 40, hidden=True)
    wb.save(f"{base_path}source/temp.xlsx")


def input_info(info_list):
    base_path = "../source/"
    wb = load_workbook(f"{base_path}source/temp.xlsx")
    sheet = wb["order"]
    info_row = [2, 3, 4, 6, 7]
    for i in range(len(info_row)):
        sheet.cell(column=9, row=info_row[i]).value = info_list[i]
    wb.save(f"{base_path}source/temp.xlsx")


def input_rule(r):
    base_path = "../source/"
    wb = load_workbook(f"{base_path}source/temp.xlsx")
    sheet = wb["order"]
    sheet.cell(column=2, row=9).value = r
    wb.save(f"{base_path}source/temp.xlsx")


def get_price():
    base_path = "../source/"
    wb = load_workbook(f"{base_path}價格表.xlsx")
    sheet = wb["price"]

    price_list = []
    col = 2
    while sheet.cell(column=col, row=2).value is not None:
        r = 2
        temp = []
        while sheet.cell(column=col, row=r).value is not None:
            temp.append(sheet.cell(column=col, row=r).value)
            r += 1
        price_list.append(temp)
        col += 1

    return price_list


def input_function(sizes, col, row):
    base_path = "../source/"
    wb = load_workbook(f"{base_path}source/temp.xlsx")
    sheet = wb["order"]
    letter = [['C', 'D', 'E', 'F', 'G'], ['G', 'H', 'I', 'J', 'K'], ['K', 'L', 'M', 'N', 'O']]
    for c in range(col):
        s = sizes[c].split("-")
        num = int(s[1]) - int(s[0]) + 1
        for r in range(row + 1):
            sheet.cell(column=5 + 4 * c, row=r + 11).value = f"=({letter[c][1]}{r + 11}*{num})"
            sheet.cell(column=6 + 4 * c, row=r + 11).value = f"=PRODUCT({letter[c][0]}{r + 11}, {letter[c][2]}{r + 11})"
        # sub
        sheet.cell(column=4 + 4 * c, row=41).value = f"=SUM({letter[c][1]}{11}:{letter[c][1]}{11 + row})"
        sheet.cell(column=5 + 4 * c, row=41).value = f"=SUM({letter[c][2]}{11}:{letter[c][2]}{11 + row})"
        sheet.cell(column=6 + 4 * c, row=41).value = f"=SUM({letter[c][3]}{11}:{letter[c][3]}{11 + row})"
    # TOTAL
    if col == 3:
        for r in range(row + 1):
            sheet.cell(column=3 + 4 * col,
                       row=r + 11).value = f"=SUM({letter[0][3]}{r + 11}, {letter[1][3]}{r + 11}, {letter[2][3]}{r + 11})"
    elif col == 2:
        for r in range(row + 1):
            sheet.cell(column=3 + 4 * col, row=r + 11).value = f"=SUM({letter[0][3]}{r + 11}, {letter[1][3]}{r + 11})"
    elif col == 1:
        for r in range(row + 1):
            sheet.cell(column=3 + 4 * col, row=r + 11).value = f"=SUM({letter[0][3]}{r + 11})"
    sheet.cell(column=3 + 4 * col, row=41).value = f"=SUM({letter[col - 1][4]}{11}:{letter[col - 1][4]}{11 + row})"
    wb.save(f"{base_path}source/temp.xlsx")


def make_img_small(series, code):
    base_path = f"../source/Image/{series}/"
    # 開資料夾
    create_path(f"{base_path}small")

    for c in code:
        img = Image.open(f"{base_path}{c}.jpg")
        img = img.resize((120, 120))
        img.save(f"{base_path}small/{c}.jpg")


def input_img(rows, series, code):
    base_path = "../source/"
    wb = load_workbook(f"{base_path}source/temp.xlsx")
    sheet = wb["order"]
    # 調整列高
    for r in range(rows + 1):
        sheet.row_dimensions[11 + r].height = 120 * 0.75
    # 置入圖片
    i = 11
    for c in code:
        img = drawing.image.Image(f"{base_path}Image/{series}/small/{c}.jpg")
        sheet.add_image(img, f"A{i}")
        i += 1

    wb.save(f"{base_path}source/temp.xlsx")
    # wb.save(f"{base_path}source/temp_{code[0]}.xlsx")


def input_price(price_list, sl_spread, exchange_rate, series, sl):
    base_path = "../source/"
    wb = load_workbook(f"{base_path}source/temp.xlsx")
    sheet = wb["order"]
    create_path(f"{base_path}/{series}")

    for level_spread in gv.level_list:
        for country_spread in gv.country_spread:
            create_path(f"{base_path}/{series}/{country_spread}")
            for size in range(len(price_list)):
                for r in range(len(price_list[size])):
                    # 3 7 11
                    if country_spread == "RMB":
                        sheet.cell(column=3 + size * 4, row=r + 11).value = \
                            round(price_list[size][r] + sl_spread + (level_spread - 2) + gv.country_spread[
                                country_spread], 1)
                    else:
                        sheet.cell(column=3 + size * 4, row=r + 11).value = \
                            round((price_list[size][r] + sl_spread + (level_spread - 2) + gv.country_spread[
                                country_spread]) / exchange_rate, 2)

            create_path(f"{base_path}/{series}/{country_spread}/{level_spread}")
            wb.save(f"{base_path}/{series}/{country_spread}/{level_spread}/Barbieliya_{series}_{sl}_order.xlsx")


# ------------order_tran-------------
def tran_choose_original(path):
    if not os.path.isdir(path):
        return "找不到連結資料夾"
    sl_kind = []
    for f in os.listdir(path):
        if f.split("_")[-2] not in sl_kind:
            sl_kind.append(f.split("_")[-2])
    base_path = "../source/"
    wb = load_workbook(f"{base_path}source/basic_ship_{len(sl_kind)}.xlsx")
    sheet = wb["order"]
    for sl in range(len(sl_kind)):
        sheet.cell(row=3, column=3*sl+3).value = sl_kind[sl]
    wb.save(f"{base_path}source/temp_s.xlsx")
    return len(sl_kind)


def tran_set_code(path, sl_kind):
    wb = load_workbook(f"{path}/{os.listdir(path)[0]}")
    sheet = wb["order"]
    size = []
    i = 1
    while True:
        if sheet.cell(row=10, column=i).value is None:
            break
        elif sheet.cell(row=10, column=i).value[-8:] == "Subtotal":
            size.append(sheet.cell(row=10, column=i).value.split(")")[0][1:])
        i += 1
    while len(size) < 3:
        size.append("")
    code = []
    i = 11
    while True:
        if sheet.cell(row=i, column=2).value is None:
            break
        else:
            code.append(sheet.cell(row=i, column=2).value)
        i += 1
    base_path = "../source/"
    wb = load_workbook(f"{base_path}source/temp_s.xlsx")
    sheet = wb["order"]
    for i in range(sl_kind):
        sheet.cell(row=4, column=i * 3 + 3).value = size[0]
        sheet.cell(row=4, column=i * 3 + 4).value = size[1]
        sheet.cell(row=4, column=i * 3 + 5).value = size[2]
    for i in range(len(code)):
        sheet.cell(row=5 + i, column=2).value = code[i]
    sheet.row_dimensions.group(5 + len(code), 34, hidden=True)
    wb.save(f"{base_path}source/temp_s.xlsx")
    return code, size


def tran_set_img(code_list, path):
    base_path = "../source/"
    wb = load_workbook(f"{base_path}source/temp_s.xlsx")
    sheet = wb["order"]
    # 調整列高
    for r in range(len(code_list)):
        sheet.row_dimensions[5 + r].height = 120 * 0.75
    # 置入圖片
    i = 5
    for c in code_list:
        img = drawing.image.Image(f"{base_path}Image/{os.listdir(path)[0].split('_')[-3]}/small/{c}.jpg")
        sheet.add_image(img, f"A{i}")
        i += 1

    wb.save(f"{base_path}source/temp_s.xlsx")


def tran_put_lot(path, code_list, size_list):
    base_path = "../source/"
    create_path(f"{base_path}訂單生成結果")

    src_dict = {}
    for file in os.listdir(path):
        f_name = file.split("Barbieliya")[0]
        if f_name not in src_dict:
            src_dict[f_name] = {}
        source = load_workbook(f"{path}/{file}")
        src = source["order"]
        temp = []
        for i in range(len(size_list)):
            temp_c = []
            for c in range(len(code_list)):
                temp_c.append(src.cell(row=11 + c, column=i * 4 + 4).value)
            temp.append(temp_c)
        src_dict[f_name][file.split('_')[-2]] = temp

    for name in src_dict:
        wb = load_workbook(f"{base_path}source/temp_s.xlsx")
        sheet = wb["order"]
        sl = 0
        while True:
            if sheet.cell(row=3, column=3+3*sl).value is not None:
                if sheet.cell(row=3, column=3+3*sl).value in src_dict[name]:
                    data = src_dict[name][sheet.cell(row=3, column=3 + 3 * sl).value]
                    for c in range(len(data)):
                        for r in range(len(data[c])):
                            sheet.cell(row=5 + r, column=3 + 3 * sl + c).value = data[c][r]
            else:
                break
            sl += 1
        wb.save(f"{base_path}訂單生成結果/{name}出貨單.xlsx")